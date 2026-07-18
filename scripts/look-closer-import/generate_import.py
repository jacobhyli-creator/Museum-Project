#!/usr/bin/env python3
"""
generate_import.py — build the "Look Closer" guided-looking import.

Reads the authored workbook (Look_Closer_Guided_Looking_Content.xlsx, sheet
"Guided Looking Data") plus the local artwork catalog
(src/data/artworks.generated.json) to validate that every workbook row matches
exactly one artwork, then emits three files into this directory:

  * verify_lookcloser_match.sql  — read-only pre-flight (match_count per row).
  * import_lookcloser.sql        — idempotent transaction writing ONLY the two
                                   new tables (guided_looking_sets +
                                   guided_looking_hotspots).
  * README.md                    — order, matching rules, re-run safety, report.

No live database access — this reads a local xlsx + JSON and writes SQL text.
It NEVER touches explanations, images, pairings, audio, artwork metadata,
routing, or any user/session data.

Matching: normalized artist + title (lowercased; curly quotes / en/em dashes
folded to ASCII; whitespace collapsed; trailing "(YYYY)" stripped from the
title). Where the catalog holds several works with the same normalized
artist+title, the row is disambiguated by year. Nothing is guessed — a row that
does not match exactly one artwork is reported and NOT emitted.

Usage:
    python3 scripts/look-closer-import/generate_import.py
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "openpyxl is required: pip install openpyxl\n"
    )
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
APP_ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
PROJECT_ROOT = os.path.abspath(os.path.join(APP_ROOT, ".."))

WORKBOOK = os.path.join(
    PROJECT_ROOT, "Look Closer Feature Explanations.xlsx"
)
CATALOG = os.path.join(APP_ROOT, "src", "data", "artworks.generated.json")
SHEET = "Guided Looking Data"

OUT_VERIFY = os.path.join(HERE, "verify_lookcloser_match.sql")
OUT_IMPORT = os.path.join(HERE, "import_lookcloser.sql")
OUT_README = os.path.join(HERE, "README.md")

# Column indices (0-based) in the "Guided Looking Data" sheet.
COL_ID = 0
COL_TITLE = 1
COL_ARTIST = 2
COL_ROOM = 3
COL_WHOLE = 5
COL_STEPBACK = 6
# Hotspot blocks: (title, what, why, question, x, y) start columns.
HOTSPOTS = [
    {"title": 7, "what": 8, "why": 9, "question": 10, "x": 11, "y": 12},
    {"title": 13, "what": 14, "why": 15, "question": 16, "x": 17, "y": 18},
    {"title": 19, "what": 20, "why": 21, "question": 22, "x": 23, "y": 24},
]
COL_MAIN_SOURCE = 25
COL_ADDL_SOURCE = 26
COL_SOURCE_NOTES = 27
COL_CONFIDENCE = 28
COL_HUMAN_REVIEWED = 29
COL_ADMIN_NOTES = 30


# ---------------------------------------------------------------------------
# Normalization — must mirror the SQL side exactly.
#   lower( translate(x, '“”‘’–—', '""''--') ) with whitespace collapsed, and
#   the trailing "(YYYY)" removed from titles.
# ---------------------------------------------------------------------------
_PUNCT = str.maketrans({
    "\u201c": '"', "\u201d": '"',   # “ ”
    "\u2018": "'", "\u2019": "'",   # ‘ ’
    "\u2013": "-", "\u2014": "-",   # – —
})
# A trailing year on a title, written either "(YYYY)" or ", YYYY".
_YEAR_SUFFIX = re.compile(r"\s*[\(,]\s*(?:c\.?\s*)?(\d{4})(?:[\-/]\d{2,4})?\)?\s*$")


def norm(value):
    s = "" if value is None else str(value)
    s = s.translate(_PUNCT)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_title(value):
    return _YEAR_SUFFIX.sub("", norm(value)).strip()


def title_year(value):
    """Extract a trailing year from a title ("(1971)" / ", 1983"), else None."""
    m = _YEAR_SUFFIX.search(norm(value))
    return m.group(1) if m else None


def accession(*values):
    """Extract the SFMOMA accession code from an /artwork/<CODE>/ URL."""
    for v in values:
        if not v:
            continue
        m = re.search(r"/artwork/([^/]+)/", str(v))
        if m:
            return m.group(1).replace(" ", "").upper()
    return None


def year_str(value):
    if value is None:
        return ""
    s = str(value).strip()
    m = re.search(r"\d{4}", s)
    return m.group(0) if m else s


def dollar_quote(text):
    """Wrap text in a Postgres dollar-quote tag that cannot collide with it."""
    text = "" if text is None else str(text)
    base = "s"
    tag = f"${base}$"
    n = 0
    while tag in text:
        n += 1
        tag = f"${base}{n}$"
    return f"{tag}{text}{tag}"


def cell(value):
    """Trim a spreadsheet cell to a clean string, or None if empty."""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def as_number(value):
    """Return a numeric literal string for x/y, or None if unusable/out of range."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        n = float(value)
    except (TypeError, ValueError):
        return None
    if n < 0 or n > 100:
        return None
    # Keep integers tidy (61 not 61.0); otherwise preserve the decimal.
    if n == int(n):
        return str(int(n))
    return repr(n)


def load_catalog():
    with open(CATALOG, encoding="utf-8") as f:
        arr = json.load(f)
    return arr if isinstance(arr, list) else arr.get("artworks", [])


def build_catalog_index(catalog):
    """Build two lookups over the catalog:
      by_title: (n_artist, n_title) -> [entry, ...]  (dups share a key)
      by_acc:   accession_code       -> [entry, ...]
    where entry = {code, year}. `code` is the stable external id (A001) that
    equals artworks.code in the DB, so the emitted SQL matches on it directly.
    The importer resolves by normalized artist+title (year-filtered for
    duplicates), falling back to the accession code for the title mismatches.
    """
    by_title = {}
    by_acc = {}
    for a in catalog:
        entry = {"code": a.get("id"), "year": year_str(a.get("year"))}
        key = (norm(a.get("artist")), norm_title(a.get("title")))
        by_title.setdefault(key, []).append(entry)
        acc = accession(a.get("preferredImageSourcePage"))
        if acc:
            by_acc.setdefault(acc, []).append(entry)
    return by_title, by_acc


def main():
    if not os.path.exists(WORKBOOK):
        sys.stderr.write(f"Workbook not found: {WORKBOOK}\n")
        sys.exit(1)
    if not os.path.exists(CATALOG):
        sys.stderr.write(f"Catalog not found: {CATALOG}\n")
        sys.exit(1)

    catalog = load_catalog()
    by_title, by_acc = build_catalog_index(catalog)

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[COL_TITLE] is None and r[COL_ARTIST] is None:
            continue  # skip fully blank spacer rows
        rows.append(r)

    parsed = []
    report = {
        "rows_read": len(rows),
        "matched": 0,
        "unmatched": [],
        "ambiguous": [],
        "sets": 0,
        "hotspots": 0,
        "hotspots_missing_coord": 0,
        "needs_review": 0,
        "published_sets": 0,
    }

    for r in rows:
        lc_id = cell(r[COL_ID]) or "(no id)"
        artist_raw = cell(r[COL_ARTIST])
        title_raw = cell(r[COL_TITLE])
        n_artist = norm(artist_raw)
        n_title = norm_title(title_raw)
        t_year = title_year(title_raw)
        row_acc = accession(r[COL_MAIN_SOURCE], r[COL_ADDL_SOURCE])

        # Layered matching, resolving to a single catalog code (== artworks.code):
        #   1. normalized artist+title (unique),
        #   2. same, disambiguated by a year in the title, for duplicate titles,
        #   3. accession code from the row's Main/Additional source, for the few
        #      title mismatches (e.g. Polke "Untitled" vs "Ohne Titel (Untitled)").
        code = None
        candidates = by_title.get((n_artist, n_title), [])
        if len(candidates) == 1:
            code = candidates[0]["code"]
        elif len(candidates) > 1 and t_year:
            year_hits = [c for c in candidates if c["year"] == t_year]
            if len(year_hits) == 1:
                code = year_hits[0]["code"]
        if code is None and row_acc and len(by_acc.get(row_acc, [])) == 1:
            code = by_acc[row_acc][0]["code"]

        if code is None:
            if len(candidates) == 0 and not row_acc:
                report["unmatched"].append(
                    {"id": lc_id, "artist": artist_raw, "title": title_raw}
                )
            else:
                report["ambiguous"].append(
                    {"id": lc_id, "artist": artist_raw, "title": title_raw,
                     "years": [c["year"] for c in candidates]}
                )
            continue

        report["matched"] += 1

        confidence = cell(r[COL_CONFIDENCE]) or ""
        publishable = confidence in ("High", "Medium")
        review_status = "approved" if publishable else "needs_review"
        is_published = publishable
        if not publishable:
            report["needs_review"] += 1
        else:
            report["published_sets"] += 1

        hotspots = []
        for i, h in enumerate(HOTSPOTS, start=1):
            htitle = cell(r[h["title"]])
            what = cell(r[h["what"]])
            why = cell(r[h["why"]])
            question = cell(r[h["question"]])
            if not any([htitle, what, why, question]):
                continue
            x = as_number(r[h["x"]])
            y = as_number(r[h["y"]])
            has_coords = x is not None and y is not None
            if not has_coords:
                report["hotspots_missing_coord"] += 1
            # A hotspot is published only when the set is published AND it has
            # both coordinates.
            hs_published = bool(is_published and has_coords)
            hotspots.append({
                "number": i,
                "title": htitle,
                "what": what,
                "why": why,
                "question": question,
                "x": x,
                "y": y,
                "is_published": hs_published,
            })
            report["hotspots"] += 1

        parsed.append({
            "lc_id": lc_id,
            "room": cell(r[COL_ROOM]),
            "artist": artist_raw,
            "title": title_raw,
            "code": code,
            "whole": cell(r[COL_WHOLE]),
            "stepback": cell(r[COL_STEPBACK]),
            "main_source": cell(r[COL_MAIN_SOURCE]),
            "addl_source": cell(r[COL_ADDL_SOURCE]),
            "source_notes": cell(r[COL_SOURCE_NOTES]),
            "confidence": confidence or None,
            "human_reviewed": (cell(r[COL_HUMAN_REVIEWED]) or "").lower()
                              in ("yes", "true", "y", "1"),
            "admin_notes": cell(r[COL_ADMIN_NOTES]),
            "review_status": review_status,
            "is_published": is_published,
            "hotspots": hotspots,
        })
        report["sets"] += 1

    write_verify(parsed)
    write_import(parsed)
    write_readme(report)
    print_report(report)

    # Non-zero exit if anything didn't cleanly match, so CI/humans notice.
    if report["unmatched"] or report["ambiguous"]:
        sys.exit(2)


# ---------------------------------------------------------------------------
# SQL emitters
#
# Each workbook row was resolved in Python to a single catalog `code` (the
# stable external id, e.g. "A002"), which equals public.artworks.code. The SQL
# therefore matches on `a.code`, sidestepping every artist/title/year/quote
# quirk. The verify query counts artworks per code (must be exactly 1).
# ---------------------------------------------------------------------------
def write_verify(parsed):
    lines = []
    lines.append(
        "-- ============================================================================"
    )
    lines.append(
        "-- verify_lookcloser_match.sql  (PRE-FLIGHT — read-only, makes NO changes)"
    )
    lines.append(
        "-- Run in the Supabase SQL Editor BEFORE import_lookcloser.sql to confirm every"
    )
    lines.append(
        "-- guided-looking row matches exactly one artwork by its code. Any row with"
    )
    lines.append(
        "-- match_count <> 1 must be resolved before importing (0 = no such code)."
    )
    lines.append(
        "-- ============================================================================"
    )
    lines.append("with wanted (code) as (")
    lines.append("  values")
    values = [f"  ({dollar_quote(p['code'])})" for p in parsed]
    lines.append(",\n".join(values))
    lines.append(")")
    lines.append("select w.code, count(a.id) as match_count")
    lines.append("from wanted w")
    lines.append("left join public.artworks a on a.code = w.code")
    lines.append("group by w.code")
    lines.append("order by match_count, w.code;")
    lines.append("-- Expect: every row match_count = 1. Rows with 0 need attention.")
    with open(OUT_VERIFY, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def write_import(parsed):
    L = []
    L.append(
        "-- ============================================================================"
    )
    L.append("-- import_lookcloser.sql  (generated — do not hand-edit)")
    L.append("-- \"Look Closer\" guided-looking content from")
    L.append("--   Look_Closer_Guided_Looking_Content.xlsx  (sheet: Guided Looking Data)")
    L.append("-- Writes ONLY public.guided_looking_sets and public.guided_looking_hotspots.")
    L.append("-- Never touches explanations, images, pairings, audio, artwork metadata,")
    L.append("-- routing, or any user/session data.")
    L.append("-- Matched by stable artwork code (a.code, e.g. 'A002').")
    L.append("-- High/Medium confidence -> review_status='approved', is_published=true so")
    L.append("-- matched sets appear to visitors immediately (still editable in the admin).")
    L.append("-- Low/blank confidence -> review_status='needs_review', unpublished.")
    L.append("-- A hotspot is published only when the set is published AND it has both")
    L.append("-- coordinates. Run in the Supabase SQL Editor AFTER migration 0015.")
    L.append(f"-- {len(parsed)} guided-looking sets. Idempotent: re-running updates in place")
    L.append("-- (unique artwork_id; unique (artwork_id, hotspot_number)).")
    L.append(
        "-- ============================================================================"
    )
    L.append("begin;")
    L.append("")

    for p in parsed:
        room = p["room"] or ""
        L.append(f"-- {room}  {p['artist']} - {p['title']}  [{p['lc_id']} -> {p['code']}]")
        # --- upsert the set ---
        L.append("with target as (")
        L.append(f"  select a.id as artwork_id from public.artworks a where a.code = {dollar_quote(p['code'])}")
        L.append("), up_set as (")
        L.append("  insert into public.guided_looking_sets")
        L.append("    (artwork_id, whole_artwork_prompt, step_back_reflection,")
        L.append("     main_source_used, additional_source_used, source_notes,")
        L.append("     confidence, human_reviewed, admin_notes,")
        L.append("     review_status, is_published)")
        L.append("  select target.artwork_id,")
        L.append(f"    {sql_val(p['whole'])}, {sql_val(p['stepback'])},")
        L.append(f"    {sql_val(p['main_source'])}, {sql_val(p['addl_source'])}, {sql_val(p['source_notes'])},")
        L.append(f"    {sql_val(p['confidence'])}, {str(p['human_reviewed']).lower()}, {sql_val(p['admin_notes'])},")
        L.append(f"    {dollar_quote(p['review_status'])}, {str(p['is_published']).lower()}")
        L.append("  from target")
        L.append("  on conflict (artwork_id) do update set")
        L.append("    whole_artwork_prompt   = excluded.whole_artwork_prompt,")
        L.append("    step_back_reflection   = excluded.step_back_reflection,")
        L.append("    main_source_used       = excluded.main_source_used,")
        L.append("    additional_source_used = excluded.additional_source_used,")
        L.append("    source_notes           = excluded.source_notes,")
        L.append("    confidence             = excluded.confidence,")
        L.append("    human_reviewed         = excluded.human_reviewed,")
        L.append("    admin_notes            = excluded.admin_notes,")
        L.append("    review_status          = excluded.review_status,")
        L.append("    is_published           = excluded.is_published,")
        L.append("    updated_at             = now()")
        L.append("  returning id, artwork_id")
        L.append(")")

        if not p["hotspots"]:
            # No hotspots — the set upsert still needs to execute.
            L.append("select 1 from up_set;")
            L.append("")
            continue

        # --- upsert each hotspot, keyed off the set produced above ---
        hs_selects = []
        for h in p["hotspots"]:
            x = h["x"] if h["x"] is not None else "null"
            y = h["y"] if h["y"] is not None else "null"
            hs_selects.append(
                "  select up_set.id, up_set.artwork_id, "
                f"{h['number']}::int,\n"
                f"    {sql_val(h['title'])}, {sql_val(h['what'])},\n"
                f"    {sql_val(h['why'])}, {sql_val(h['question'])},\n"
                f"    {x}, {y}, {str(h['is_published']).lower()}\n"
                "  from up_set"
            )
        L.append("insert into public.guided_looking_hotspots")
        L.append("  (set_id, artwork_id, hotspot_number, title, what_to_look_at,")
        L.append("   why_it_matters, visitor_question, x_coordinate, y_coordinate, is_published)")
        L.append("\n  union all\n".join(hs_selects))
        L.append("on conflict (artwork_id, hotspot_number) do update set")
        L.append("  set_id           = excluded.set_id,")
        L.append("  title            = excluded.title,")
        L.append("  what_to_look_at  = excluded.what_to_look_at,")
        L.append("  why_it_matters   = excluded.why_it_matters,")
        L.append("  visitor_question = excluded.visitor_question,")
        L.append("  x_coordinate     = excluded.x_coordinate,")
        L.append("  y_coordinate     = excluded.y_coordinate,")
        L.append("  is_published     = excluded.is_published,")
        L.append("  updated_at       = now();")
        L.append("")

    L.append("commit;")
    with open(OUT_IMPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")


def sql_val(v):
    if v is None:
        return "null"
    return dollar_quote(v)


def write_readme(report):
    txt = f"""# Look Closer guided-looking import

Source: `Look_Closer_Guided_Looking_Content.xlsx` (sheet **Guided Looking Data**,
{report['rows_read']} rows).

Writes **only** `public.guided_looking_sets` and `public.guided_looking_hotspots`.
Never touches explanations, images, pairings, audio, artwork metadata, routing, or
user/session data.

## Order

1. Apply migration **`0015_look_closer.sql`** (creates the two tables + RLS).
2. Run **`verify_lookcloser_match.sql`** in the Supabase SQL Editor. It is read-only.
   Confirm every row shows `match_count = 1`. If any row is `0` (no artwork) or `2+`
   (ambiguous), fix the artwork title/artist/year first — do **not** import until clean.
3. Run **`import_lookcloser.sql`**. It upserts the guided-looking sets and their
   hotspots. High/Medium-confidence sets are `review_status='approved'`,
   `is_published=true`, so they appear to visitors immediately (and stay editable in
   the admin). Low/blank-confidence sets import as `needs_review` and unpublished.

## Matching

- Normalized **artist + title** (lowercased; curly quotes and en/em dashes folded to
  ASCII; whitespace collapsed; a trailing `(YYYY)` stripped from the title).
- Where the catalog holds several works with the same normalized artist+title, the row
  is disambiguated by **year**.
- Nothing is guessed. A spreadsheet row that doesn't match exactly one artwork is
  reported below and simply not inserted.

## Publishing rules

- A **set** is publishable only when confidence is `High` or `Medium`.
- A **hotspot** is `is_published=true` only when its set is published AND it has both
  X and Y coordinates.
- The public app shows Look Closer only when the set is approved + published AND at
  least one hotspot is published.

## Re-running

`import_lookcloser.sql` is idempotent (`on conflict (artwork_id)` for sets;
`on conflict (artwork_id, hotspot_number)` for hotspots, both `do update`), so it is
safe to re-run after editing the spreadsheet and regenerating.

## Report (last generated)

- Rows read: **{report['rows_read']}**
- Matched to an artwork: **{report['matched']}**
- Unmatched (no artwork): **{len(report['unmatched'])}**
- Ambiguous (2+ artworks): **{len(report['ambiguous'])}**
- Sets emitted: **{report['sets']}**  (published: {report['published_sets']}, needs_review: {report['needs_review']})
- Hotspots emitted: **{report['hotspots']}**  (missing coordinates: {report['hotspots_missing_coord']})
"""
    if report["unmatched"]:
        txt += "\n### Unmatched rows\n\n"
        for u in report["unmatched"]:
            txt += f"- `{u['id']}` {u['artist']} — {u['title']}\n"
    if report["ambiguous"]:
        txt += "\n### Ambiguous rows (need year disambiguation)\n\n"
        for u in report["ambiguous"]:
            txt += f"- `{u['id']}` {u['artist']} — {u['title']} (catalog years: {', '.join(u['years'])})\n"
    with open(OUT_README, "w", encoding="utf-8") as f:
        f.write(txt)


def print_report(report):
    print("Look Closer import — generation report")
    print(f"  Rows read:              {report['rows_read']}")
    print(f"  Matched to an artwork:  {report['matched']}")
    print(f"  Unmatched (no artwork): {len(report['unmatched'])}")
    print(f"  Ambiguous (2+ works):   {len(report['ambiguous'])}")
    print(f"  Sets emitted:           {report['sets']} "
          f"(published {report['published_sets']}, needs_review {report['needs_review']})")
    print(f"  Hotspots emitted:       {report['hotspots']} "
          f"(missing coords {report['hotspots_missing_coord']})")
    for u in report["unmatched"]:
        print(f"    UNMATCHED  {u['id']}  {u['artist']} - {u['title']}")
    for u in report["ambiguous"]:
        print(f"    AMBIGUOUS  {u['id']}  {u['artist']} - {u['title']} "
              f"(years: {', '.join(u['years'])})")


if __name__ == "__main__":
    main()
